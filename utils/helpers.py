from pathlib import Path
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import sqlite3, zipfile, os
from openpyxl import Workbook, load_workbook
from PyQt6 import QtWidgets

DB = Path(__file__).parent.parent / "database" / "graduates.db"
INSTITUTE_NAME = "Sukkur IBA University Kandhkot Campus"


def export_to_excel(dest):
    """Export all graduate records to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.append([
        "id", "name", "student_id", "department", "batch", "bio", "skills",
        "internship", "degree", "certificates", "profile_img", "verified"
    ])
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT * FROM graduates")
    for row in cur.fetchall():
        ws.append(list(row))
    conn.close()
    wb.save(dest)


def import_from_excel(src_path):
    """Import graduate records from Excel into the database."""
    try:
        wb = load_workbook(src_path)
        ws = wb.active
        conn = sqlite3.connect(DB)
        cur = conn.cursor()

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_count = len(headers)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < col_count:
                continue
            # Skip if name or ID is missing
            if not row[1] or not row[2]:
                continue
            cur.execute("""
                INSERT OR REPLACE INTO graduates 
                (id, name, student_id, department, batch, bio, skills, internship, degree, certificates, profile_img, verified)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, row)
        conn.commit()
        conn.close()

        QtWidgets.QMessageBox.information(None, "Import Successful", "Graduates imported from Excel successfully.")
    except Exception as e:
        QtWidgets.QMessageBox.critical(None, "Import Error", str(e))


def export_profile_pdf(record, dest):
    """Export a single graduate's profile as PDF."""
    c = canvas.Canvas(dest, pagesize=letter)
    x = 72
    y = 720
    c.setFont("Helvetica-Bold", 18)
    c.drawString(x, y, INSTITUTE_NAME.upper())
    y -= 28
    c.setFont("Helvetica-Bold", 16)
    c.drawString(x, y, record[1])
    y -= 24
    c.setFont("Helvetica", 11)
    lines = [
        f"ID: {record[2]}",
        f"Department: {record[3]}",
        f"Batch: {record[4]}",
        "",
        "Bio:",
        record[5] or ""
    ]
    for line in lines:
        c.drawString(x, y, str(line))
        y -= 14
        if y < 72:
            c.showPage()
            y = 720
    c.save()


def zip_attachments(out_path, attachments_folder):
    """Compress attachments folder into a single ZIP archive."""
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(attachments_folder):
            for f in files:
                full = os.path.join(root, f)
                z.write(full, arcname=os.path.relpath(full, attachments_folder))
